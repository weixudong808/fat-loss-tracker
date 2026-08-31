#!/usr/bin/env python3
"""fat-loss-tracker 公共库：配置、存储后端（本地 xlsx / 飞书电子表格）、
档案、阶梯减重、渐进超负荷。
被 record.py / profile.py / plan.py / storage.py / bootstrap.py 复用，不单独执行业务。

存储后端：
- local：本地 xlsx 工作簿（openpyxl），零安装零授权，首次运行自动建表
- feishu：飞书电子表格（本机 lark-cli），需 config.json 配 spreadsheet_url
后端选择：load_config() 自动完成——已配置直接用；未配置则本地优先探测，飞书兜底。
"""
import csv
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import date, datetime, timedelta
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
CONFIG_PATH = SCRIPT_DIR / "config.json"

# 全新环境的空配置模板：backend 留空 → ensure_backend 自动探测（本地 xlsx 优先）
DEFAULT_CONFIG = {
    "backend": "",
    "local_path": "",
    "spreadsheet_url": "",
    "sheet_ids": {"训练记录": "", "体重记录": "", "饮食记录": "", "用户档案": ""},
    "cron_job_ids": {},
}
# 数据表放技能目录外（SKILL_DIR 旁的 减脂数据/）：千问正式注册禁止技能包内含二进制文件
# （xlsx 会被安全审核拒），且更新技能（删目录重装）时数据不受影响。老用户 config.json 的 local_path 优先于本默认。
DEFAULT_XLSX = SKILL_DIR.parent / "减脂数据" / "减脂追踪数据.xlsx"

# ---------------------------------------------------------------- 常量
# 用户档案表列顺序（单行配置：第1行表头，第2行数据）
PROFILE_COLS = [
    "创建日期", "性别", "年龄", "身高cm",
    "初始体重kg", "当前体重kg", "目标体重kg", "体脂率%",
    "活动量", "每周训练天数", "训练日安排", "训练偏好",
    "当前档位kg", "减重路径",
    "碳水g", "蛋白g", "脂肪g",
    "AI称呼", "回复风格", "记忆备注", "用户昵称",
]
# 训练记录单表承载实际记录与滚动计划：状态=已完成（实际）/待完成（自动生成的下次计划）/已跳过（被新记录替代）
TRAIN_COLS = ["日期", "训练主题", "动作名称", "组数", "次数", "重量kg", "备注", "状态"]
WEIGHT_COLS = ["日期", "体重kg", "体脂率%", "备注"]
DIET_COLS = ["日期", "餐次", "食物内容", "备注"]
PLAN_COLS = TRAIN_COLS  # 兼容旧引用
PLAN_STATUSES = ["待完成", "已完成", "已跳过"]
REP_LADDER = [10, 12, 15]  # 渐进超负荷次数阶梯
WEIGHT_STEP_KG = 5         # 阶梯式减重：每 5kg 一档
# 本地工作簿四张工作表的表头
SHEET_HEADERS = {
    "训练记录": TRAIN_COLS,
    "体重记录": WEIGHT_COLS,
    "饮食记录": DIET_COLS,
    "用户档案": PROFILE_COLS,
}

# ---------------------------------------------------------------- 配置 / 后端探测
def save_config(config):
    """原子写回 config.json。"""
    tmp = CONFIG_PATH.with_suffix(".json.tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
    os.replace(tmp, CONFIG_PATH)


def load_config():
    """读配置并确保存储后端可用（未配置时自动探测并初始化，结果写回 config.json）。"""
    if not CONFIG_PATH.exists():
        # 全新环境：自动落空模板再走自动探测，首次运行零配置初始化
        config = json.loads(json.dumps(DEFAULT_CONFIG))
        save_config(config)
    else:
        with open(CONFIG_PATH, encoding="utf-8") as f:
            config = json.load(f)
    return ensure_backend(config)


def ensure_backend(config):
    """返回可用的 config；必要时自动探测/初始化后端。
    1) backend 已显式指定 → 直接用（local 文件丢了自动重建空表）
    2) 老配置无 backend 但有飞书 URL → 标记 feishu（老用户无感）
    3) 全新配置 → 本地 xlsx 优先；不可用再看 lark-cli；都没有则报错给指引
    """
    backend = config.get("backend")
    if backend == "local":
        path = _local_path(config)
        if not path.exists():
            _build_local_workbook(path)
        return config
    if backend == "feishu":
        if not config.get("spreadsheet_url"):
            die("backend=feishu 但 spreadsheet_url 为空：请在 config.json 填入飞书表格链接，"
                "或删掉 backend 字段重新自动探测")
        return config
    # 老配置兼容：已有飞书链接直接走飞书
    if config.get("spreadsheet_url"):
        config["backend"] = "feishu"
        save_config(config)
        return config
    # 自动探测：本地优先
    if _openpyxl_available() and _dir_writable(DEFAULT_XLSX.parent):
        return bootstrap_local(config, DEFAULT_XLSX)
    if shutil.which("lark-cli"):
        die("检测到 lark-cli 但未配置飞书表格：请在 config.json 填入 spreadsheet_url 后重试；"
            "或执行 pip install openpyxl 后删掉 backend 字段，自动启用本地表格")
    die("未找到可用存储后端：推荐执行 pip install openpyxl 启用本地表格（零配置）；"
        "或安装 lark-cli 并在 config.json 配置飞书电子表格链接")


def _openpyxl_available():
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


def _dir_writable(d):
    try:
        Path(d).mkdir(parents=True, exist_ok=True)
        with tempfile.NamedTemporaryFile(dir=str(d), delete=True):
            pass
        return True
    except Exception:
        return False


def detect_report():
    """供 bootstrap.py status：返回环境探测结果（不做任何修改）。"""
    return {
        "ok": True,
        "config_path": str(CONFIG_PATH),
        "backend_in_config": json.loads(CONFIG_PATH.read_text(encoding="utf-8")).get("backend")
        if CONFIG_PATH.exists() else None,
        "openpyxl_available": _openpyxl_available(),
        "skill_dir_writable": _dir_writable(SKILL_DIR),
        "lark_cli_path": shutil.which("lark-cli"),
        "default_local_path": str(DEFAULT_XLSX),
    }


def bootstrap_local(config, path=None):
    """启用本地后端：建 xlsx 工作簿（4 表+表头）并把 backend/local_path 写回 config。"""
    if not _openpyxl_available():
        die("本地后端需要 openpyxl，请先执行：pip install openpyxl")
    path = Path(path) if path else _local_path(config)
    if not path.is_absolute():
        path = (SCRIPT_DIR / path).resolve()
    _build_local_workbook(path)
    config["backend"] = "local"
    config["local_path"] = str(path)
    config.setdefault("spreadsheet_url", "")
    save_config(config)
    return config


def bootstrap_feishu(config, url):
    """显式启用飞书后端并记录 URL。"""
    if not url:
        die("启用飞书后端需要 spreadsheet_url")
    config["backend"] = "feishu"
    config["spreadsheet_url"] = url
    config.setdefault("local_path", "")
    save_config(config)
    return config


def storage_info(config):
    """当前数据位置（录入回执/SKILL 回复用）。"""
    if config.get("backend") == "local":
        return {"backend": "local", "file_path": str(_local_path(config)), "sheet_url": None}
    return {"backend": "feishu", "file_path": None,
            "sheet_url": config.get("spreadsheet_url")}


def _local_path(config):
    p = config.get("local_path") or str(DEFAULT_XLSX)
    p = Path(p)
    return p if p.is_absolute() else (SCRIPT_DIR / p).resolve()


# ---------------------------------------------------------------- 本地 xlsx 后端
def _build_local_workbook(path):
    """创建本地工作簿：四张工作表 + 表头行。已存在则不覆盖。"""
    import openpyxl
    path = Path(path)
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, cols in SHEET_HEADERS.items():
        ws = wb.create_sheet(name)
        ws.append(cols)
    _atomic_save(wb, path)


def _local_load_wb(path):
    import openpyxl
    return openpyxl.load_workbook(path)


def _atomic_save(wb, path):
    """先写临时文件再原子替换，避免写一半中断损坏数据。"""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=str(path.parent))
    os.close(fd)
    try:
        wb.save(tmp)
        os.replace(tmp, path)
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)


def _col_index(letter):
    from openpyxl.utils import column_index_from_string
    return column_index_from_string(str(letter).strip().replace("$", ""))


def _norm_cell(v):
    """读出值归一：datetime→YYYY-MM-DD 字符串；整数浮点→int；字符串 strip。对齐飞书版 _auto_num。"""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str):
        return v.strip()
    return v


def _local_read(config, sheet_name, last_col, max_row=500):
    path = _local_path(config)
    if not path.exists():
        return None, {"ok": False, "error": f"本地数据文件不存在: {path}"}
    wb = _local_load_wb(path)
    if sheet_name not in wb.sheetnames:
        names = wb.sheetnames
        wb.close()
        return None, {"ok": False, "error": f"工作簿缺少工作表「{sheet_name}」，现有: {names}"}
    ws = wb[sheet_name]
    max_c = _col_index(last_col)
    header, records = [], []
    last_r = min(ws.max_row or 1, max_row)
    for r in range(1, last_r + 1):
        vals = [_norm_cell(ws.cell(row=r, column=c).value) for c in range(1, max_c + 1)]
        if not any(v not in (None, "") for v in vals):
            continue
        if r == 1:
            header = [("" if v is None else str(v)).strip() for v in vals]
            continue
        rec = {h: vals[i] for i, h in enumerate(header) if h}
        rec["_row"] = r
        records.append(rec)
    wb.close()
    return header, records


def _local_append(config, sheet_name, columns, dtypes, formats, rows):
    path = _local_path(config)
    if not path.exists():
        _build_local_workbook(path)
    wb = _local_load_wb(path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {"ok": False, "error": f"工作簿缺少工作表「{sheet_name}」"}
    ws = wb[sheet_name]
    # 最后一个非空行（表头必占第 1 行），从其后追加，不重复写表头
    last = 1
    for r in range(1, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value not in (None, "")
               for c in range(1, ws.max_column + 1)):
            last = r
    header_map = {str(ws.cell(row=1, column=c).value or "").strip(): c
                  for c in range(1, ws.max_column + 1)}
    col_idx = []
    for name in columns:
        c = header_map.get(name)
        if c is None:
            wb.close()
            return {"ok": False, "error": f"工作表「{sheet_name}」无此列: {name}"}
        col_idx.append(c)
    for i, row in enumerate(rows):
        for c, val in zip(col_idx, row):
            ws.cell(row=last + 1 + i, column=c, value=val)
    _atomic_save(wb, path)
    wb.close()
    return {"ok": True, "appended": len(rows)}


def _local_set_cells(config, sheet_name, a1_range, values_2d):
    from openpyxl.utils import range_boundaries
    path = _local_path(config)
    if not path.exists():
        return {"ok": False, "error": f"本地数据文件不存在: {path}"}
    wb = _local_load_wb(path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {"ok": False, "error": f"工作簿缺少工作表「{sheet_name}」"}
    ws = wb[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(a1_range)
    for i, row in enumerate(values_2d):
        for j, v in enumerate(row):
            ws.cell(row=min_row + i, column=min_col + j,
                    value=None if v is None else v)
    _atomic_save(wb, path)
    wb.close()
    return {"ok": True, "updated_range": a1_range}


# ---------------------------------------------------------------- 飞书后端（lark-cli）
def die(msg):
    print(json.dumps({"ok": False, "error": msg}, ensure_ascii=False))
    sys.exit(1)


def lark(args, stdin_data=None, timeout=60):
    """执行 lark-cli 并解析 JSON 返回。"""
    cmd = ["lark-cli", "sheets"] + args
    result = subprocess.run(cmd, capture_output=True, text=True,
                            input=stdin_data, timeout=timeout)
    if result.returncode != 0:
        return {"ok": False, "error": (result.stderr or result.stdout).strip()[:500]}
    try:
        return json.loads(result.stdout)
    except json.JSONDecodeError:
        return {"ok": False, "error": "无法解析 lark-cli 输出", "raw": result.stdout[:500]}


def _feishu_append(config, sheet_name, columns, dtypes, formats, rows):
    payload = {"sheets": [{
        "name": sheet_name, "mode": "append", "header": False,
        "columns": columns, "dtypes": dtypes, "formats": formats, "data": rows,
    }]}
    return lark(["+table-put", "--url", config["spreadsheet_url"], "--sheets", "-"],
                stdin_data=json.dumps(payload, ensure_ascii=False))


def _auto_num(v):
    v = (v or "").strip()
    if v == "":
        return None
    try:
        f = float(v)
        return int(f) if f.is_integer() else f
    except ValueError:
        return v


def _feishu_read(config, sheet_name, last_col, max_row=500):
    """读子表，返回 (表头list, 记录list[dict])。每条记录额外带 _row（表格行号，1-based）。"""
    rng = f"A1:{last_col}{max_row}"
    r = lark(["+csv-get", "--url", config["spreadsheet_url"],
              "--sheet-name", sheet_name, "--range", rng])
    if not r.get("ok"):
        return None, r
    annotated = r.get("data", {}).get("annotated_csv", "")
    header, records = [], []
    for line in annotated.split("\n"):
        m = re.match(r"\[row=(\d+)\]\s?(.*)$", line)
        if not m:
            continue
        row_idx = int(m.group(1))
        vals = next(csv.reader([m.group(2)]))
        if not any((v or "").strip() for v in vals):
            continue
        if row_idx == 1:
            header = [v.strip() for v in vals]
            continue
        rec = {h: _auto_num(v) for h, v in zip(header, vals)}
        rec["_row"] = row_idx
        records.append(rec)
    return header, records


def _feishu_set(config, sheet_name, a1_range, values_2d):
    """按区域写值（values_2d 为行×列，None/"" 写空串）。"""
    cells = [[{"value": "" if v is None else v} for v in row] for row in values_2d]
    return lark(["+cells-set", "--url", config["spreadsheet_url"],
                 "--sheet-name", sheet_name, "--range", a1_range,
                 "--cells", json.dumps(cells, ensure_ascii=False)])


# ---------------------------------------------------------------- 存储接口（后端无关，上层只用这三个）
def append_rows(config, sheet_name, columns, dtypes, formats, rows):
    if config.get("backend") == "local":
        return _local_append(config, sheet_name, columns, dtypes, formats, rows)
    return _feishu_append(config, sheet_name, columns, dtypes, formats, rows)


def read_sheet(config, sheet_name, last_col, max_row=500):
    """读子表 → (表头list, 记录list[dict]，每条带 _row 1-based)。出错时 records 为错误 dict。"""
    if config.get("backend") == "local":
        return _local_read(config, sheet_name, last_col, max_row)
    return _feishu_read(config, sheet_name, last_col, max_row)


def set_cells(config, sheet_name, a1_range, values_2d):
    if config.get("backend") == "local":
        return _local_set_cells(config, sheet_name, a1_range, values_2d)
    return _feishu_set(config, sheet_name, a1_range, values_2d)


def out(obj, code=None):
    print(json.dumps(obj, ensure_ascii=False, indent=2))
    sys.exit(code if code is not None else (0 if obj.get("ok", True) else 1))

# ---------------------------------------------------------------- 用户档案
def get_profile(config):
    """读取档案（第2行）。无档案返回 None。"""
    _, records = read_sheet(config, "用户档案", "U")
    if isinstance(records, dict):  # 出错
        return records
    if not records:
        return None
    rec = records[0]
    rec.pop("_row", None)
    return rec


def upsert_profile(config, fields):
    """合并写回档案单行（第2行）。fields 中 None 不覆盖已有值。返回合并后的完整档案。"""
    _, existing_rows = read_sheet(config, "用户档案", "U")
    if isinstance(existing_rows, dict) and not existing_rows.get("ok", True):
        return existing_rows
    current = {k: None for k in PROFILE_COLS}
    if existing_rows:
        for k in PROFILE_COLS:
            current[k] = existing_rows[0].get(k)
    for k, v in fields.items():
        if k not in PROFILE_COLS:
            return {"ok": False, "error": f"档案无此字段: {k}，合法字段：{PROFILE_COLS}"}
        current[k] = v  # 显式传入即写入（含 None/"" = 清空）；未传入的键保持原值
    values = [[current.get(k) for k in PROFILE_COLS]]
    r = set_cells(config, "用户档案", "A2:U2", values)
    if not r.get("ok"):
        return r
    return {"ok": True, "profile": current}

# ---------------------------------------------------------------- 阶梯减重 / 营养素
def build_tiers(start_weight, target_weight, step=WEIGHT_STEP_KG):
    """初始体重 → 目标体重，每 step kg 一档（含首尾）。"""
    if start_weight is None or target_weight is None or start_weight <= target_weight:
        return []
    tiers = []
    w = start_weight
    while w > target_weight:
        w -= step
        tiers.append(round(max(w, target_weight), 1))
    if not tiers or tiers[-1] != target_weight:
        tiers.append(round(target_weight, 1))
    return tiers


def protein_multiplier(days_per_week):
    """按每周训练天数定蛋白倍数：1-2练×1 / 3-4练×1.2 / 5-6练×1.5。"""
    try:
        d = int(days_per_week)
    except (TypeError, ValueError):
        return 1.2
    if d <= 2:
        return 1.0
    if d <= 4:
        return 1.2
    return 1.5


def calc_macros(tier_weight, days_per_week, carb_mult=3.0):
    """按当前档位体重算营养素（克）。碳水×2~3（新手3起步，适应后2）；蛋白按天数；脂肪×1。"""
    if tier_weight is None:
        return {"carb_g": None, "protein_g": None, "fat_g": None}
    return {
        "carb_g": round(tier_weight * carb_mult),
        "protein_g": round(tier_weight * protein_multiplier(days_per_week)),
        "fat_g": round(tier_weight * 1.0),
    }


def advance_tier(profile, today_weight):
    """阶段联动：体重达标则进下一档。返回 (新档案fields, 进阶信息dict 或 None)。"""
    cur_tier = profile.get("当前档位kg")
    path = (profile.get("减重路径") or "")
    tiers = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", str(path))]
    if cur_tier is None or today_weight is None:
        return {}, None
    try:
        cur_tier = float(cur_tier)
    except (TypeError, ValueError):
        return {}, None
    if today_weight > cur_tier:
        return {}, None
    # 达标：找下一档
    next_tier = None
    for t in tiers:
        if t < cur_tier - 1e-9:
            next_tier = t
            break
    if next_tier is None:
        # 已到最终档
        return {}, {"reached_final": True, "tier": cur_tier}
    macros = calc_macros(next_tier, profile.get("每周训练天数"))
    fields = {"当前档位kg": next_tier, "碳水g": macros["carb_g"],
              "蛋白g": macros["protein_g"], "脂肪g": macros["fat_g"]}
    return fields, {"advanced": True, "from_kg": cur_tier, "to_kg": next_tier, "macros": macros}

# ---------------------------------------------------------------- 渐进超负荷
def next_progression(weight, reps):
    """次数 10→12→15；到15后重量+5%、次数回10。返回 (新重量, 新次数)。"""
    try:
        w = float(weight or 0)
        r = int(float(reps))
    except (TypeError, ValueError):
        return weight, reps
    nxt = next((x for x in REP_LADDER if x > r), None)
    if nxt is not None:
        return _fmt_weight(w), nxt
    new_w = round(w * 1.05, 1)
    return _fmt_weight(new_w), 10


def _fmt_weight(w):
    if w is None:
        return 0
    if abs(w - round(w)) < 0.01:
        return int(round(w))
    return round(w, 1)

# ---------------------------------------------------------------- 训练日
def parse_training_days(text):
    """'1,3,5' / '一三五' / '周一周三' → [1,3,5]（1=周一…7=周日）。"""
    if text is None:
        return []
    s = str(text)
    cn = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "日": 7, "天": 7}
    days = set()
    for ch in s:
        if ch in cn:
            days.add(cn[ch])
    for n in re.findall(r"[1-7]", s):
        days.add(int(n))
    return sorted(days)


def next_training_date(profile, after=None):
    """档案训练日安排之后的第一个训练日；未配置则返回明天。"""
    after = after or date.today()
    days = parse_training_days(profile.get("训练日安排") if profile else None)
    d = after + timedelta(days=1)
    if not days:
        return d, False
    for _ in range(7):
        if d.isoweekday() in days:  # isoweekday: Mon=1
            return d, True
        d += timedelta(days=1)
    return after + timedelta(days=1), False


def today_str():
    return datetime.now().strftime("%Y-%m-%d")
